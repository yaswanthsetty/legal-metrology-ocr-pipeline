"""
Upload Image Processing Page - Legal Metrology OCR Pipeline

Drag-and-drop image upload with multi-format support and batch processing capabilities.
Provides comprehensive results display and export functionality.
"""

import streamlit as st
import os
import time
import json
import cv2
from datetime import datetime
from pathlib import Path
import sys
from PIL import Image
import io
import base64
import zipfile
import pandas as pd
from typing import List, Dict, Any

# Add project root to path
project_root = Path(__file__).parent.parent.parent  # Go up two levels from web/pages folder
sys.path.append(str(project_root))

from live_processor import LiveProcessor
from data_refiner.refiner import DataRefiner  
from lmpc_checker.compliance_validator import ComplianceValidator
from config import EXPECTED_FIELDS

# Page configuration
st.set_page_config(
    page_title="Upload Images - Legal Metrology OCR",
    page_icon="📂",
    layout="wide"
)

# Constants
SUPPORTED_FORMATS = ['jpg', 'jpeg', 'png', 'bmp', 'webp']
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB per file
MAX_BATCH_SIZE = 20  # Maximum files per batch

def load_upload_css():
    """Load custom CSS for upload page"""
    css = """
    <style>
    .upload-zone {
        border: 2px dashed #1f4037;
        border-radius: 15px;
        padding: 2rem;
        text-align: center;
        background: linear-gradient(135deg, #f5f7fa, #c3cfe2);
        margin: 1rem 0;
        transition: all 0.3s ease;
    }
    
    .upload-zone:hover {
        border-color: #99f2c8;
        background: linear-gradient(135deg, #e3f2fd, #bbdefb);
        transform: translateY(-2px);
    }
    
    .batch-summary {
        background: #f8f9fa;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        border-left: 4px solid #1f4037;
    }
    
    .image-card {
        background: white;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        padding: 1rem;
        margin: 0.5rem 0;
        border: 1px solid #e0e0e0;
        transition: transform 0.2s ease;
    }
    
    .image-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }
    
    .progress-ring {
        display: inline-block;
        width: 40px;
        height: 40px;
        border: 3px solid #f3f3f3;
        border-top: 3px solid #1f4037;
        border-radius: 50%;
        animation: spin 1s linear infinite;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    .status-processing {
        color: #ff9800;
        font-weight: bold;
    }
    
    .status-complete {
        color: #4caf50;
        font-weight: bold;
    }
    
    .status-error {
        color: #f44336;
        font-weight: bold;
    }
    
    .export-panel {
        background: linear-gradient(135deg, #667eea, #764ba2);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .file-info {
        background: #e8f5e8;
        padding: 0.5rem;
        border-radius: 5px;
        margin: 0.25rem 0;
        border-left: 3px solid #4caf50;
    }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

def initialize_upload_session():
    """Initialize upload-specific session state"""
    if 'uploaded_files' not in st.session_state:
        st.session_state.uploaded_files = []
    
    if 'batch_results' not in st.session_state:
        st.session_state.batch_results = []
    
    if 'processing_status' not in st.session_state:
        st.session_state.processing_status = {}
    
    if 'selected_results' not in st.session_state:
        st.session_state.selected_results = []
    
    if 'current_batch_results' not in st.session_state:
        st.session_state.current_batch_results = []
    
    if 'files_processed' not in st.session_state:
        st.session_state.files_processed = set()
    
    if 'processing_completed' not in st.session_state:
        st.session_state.processing_completed = False

def validate_uploaded_file(uploaded_file) -> tuple[bool, str]:
    """Validate uploaded file"""
    # Check file extension
    file_extension = uploaded_file.name.split('.')[-1].lower()
    if file_extension not in SUPPORTED_FORMATS:
        return False, f"Unsupported format. Supported: {', '.join(SUPPORTED_FORMATS)}"
    
    # Check file size
    if uploaded_file.size > MAX_FILE_SIZE:
        return False, f"File too large. Maximum size: {MAX_FILE_SIZE // (1024*1024)}MB"
    
    # Try to open as image
    try:
        image = Image.open(uploaded_file)
        image.verify()
        return True, "Valid image file"
    except Exception as e:
        return False, f"Invalid image file: {str(e)}"

def create_file_upload_interface():
    """Create the main file upload interface"""
    st.markdown("### 📂 Upload Product Images")
    
    # Upload zone
    st.markdown("""
    <div class="upload-zone">
        <h3>📁 Drag & Drop Images Here</h3>
        <p>Or use the button below to browse files</p>
        <p><strong>Supported formats:</strong> JPG, PNG, BMP, WEBP</p>
        <p><strong>Maximum size:</strong> 10MB per file | <strong>Batch limit:</strong> 20 files</p>
    </div>
    """, unsafe_allow_html=True)
    
    # File uploader
    uploaded_files = st.file_uploader(
        "Choose image files",
        type=SUPPORTED_FORMATS,
        accept_multiple_files=True,
        help=f"Upload up to {MAX_BATCH_SIZE} images at once. Maximum {MAX_FILE_SIZE // (1024*1024)}MB per file."
    )
    
    return uploaded_files

def display_uploaded_files(uploaded_files):
    """Display uploaded files with validation status"""
    if not uploaded_files:
        # Reset processing state when no files
        st.session_state.current_batch_results = []
        st.session_state.files_processed = set()
        st.session_state.processing_completed = False
        return []
    
    # Check if files have changed
    current_file_ids = {f"{f.name}_{f.size}" for f in uploaded_files}
    if current_file_ids != st.session_state.files_processed:
        # Reset processing state for new files
        st.session_state.current_batch_results = []
        st.session_state.processing_completed = False
    
    st.markdown("### 📋 Uploaded Files")
    
    valid_files = []
    
    for i, uploaded_file in enumerate(uploaded_files):
        col1, col2, col3 = st.columns([3, 2, 1])
        
        with col1:
            # File info
            st.markdown(f"""
            <div class="file-info">
                <strong>{uploaded_file.name}</strong><br>
                Size: {uploaded_file.size / 1024:.1f} KB | Type: {uploaded_file.type}
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            # Validation status
            is_valid, message = validate_uploaded_file(uploaded_file)
            if is_valid:
                st.success(f"✅ {message}")
                valid_files.append(uploaded_file)
            else:
                st.error(f"❌ {message}")
        
        with col3:
            # Preview thumbnail
            if uploaded_file.type.startswith('image/'):
                try:
                    image = Image.open(uploaded_file)
                    st.image(image, width=100)
                    uploaded_file.seek(0)  # Reset file pointer
                except:
                    st.write("No preview")
    
    return valid_files

def process_single_image(uploaded_file, file_index: int, total_files: int) -> Dict[str, Any]:
    """Process a single uploaded image through the pipeline"""
    
    try:
        # Initialize components
        ocr_processor = LiveProcessor()
        data_refiner = DataRefiner()
        compliance_validator = ComplianceValidator()
        
        # Save uploaded file temporarily
        temp_path = f"temp_upload_{file_index}.jpg"
        
        # Convert uploaded file to image and save
        image = Image.open(uploaded_file)
        if image.mode != 'RGB':
            image = image.convert('RGB')
        image.save(temp_path)
        
        start_time = time.time()
        
        # Process through pipeline
        st.info(f"Processing {uploaded_file.name}...")
        
        # Load image for processing
        frame = cv2.imread(temp_path)
        if frame is None:
            raise ValueError(f"Could not load image from {temp_path}")
        
        st.info("Extracting text with OCR...")
        # Extract text using Surya OCR
        extracted_text = ocr_processor._extract_text_with_surya(frame)
        if not extracted_text:
            raise ValueError("No text could be extracted from the image")
        
        st.info("Structuring extracted text...")
        # Structure the text using NLP
        structured_data = ocr_processor._structure_text_with_nlp(extracted_text)
        if not structured_data:
            structured_data = ocr_processor._construct_json_from_text(extracted_text)
        
        st.info("Refining data...")
        # Create OCR result format
        ocr_result = extracted_text
        refined_data = data_refiner.refine(structured_data if structured_data else extracted_text)
        
        st.info("Validating compliance...")
        violations = compliance_validator.validate(refined_data)
        
        processing_time = time.time() - start_time
        
        # Prepare result
        result = {
            'filename': uploaded_file.name,
            'file_size': uploaded_file.size,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'method': 'File Upload',
            'processing_time': processing_time,
            'ocr_result': ocr_result,
            'refined_data': refined_data,
            'violations': violations,
            'compliance_status': 'COMPLIANT' if not violations else 'NON_COMPLIANT',
            'image_dimensions': image.size
        }
        
        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return result
        
    except Exception as e:
        # Clean up temp file on error
        temp_path = f"temp_upload_{file_index}.jpg"
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return {
            'filename': uploaded_file.name,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'method': 'File Upload',
            'error': str(e),
            'compliance_status': 'ERROR'
        }

def process_batch_images(valid_files: List) -> List[Dict[str, Any]]:
    """Process multiple images in batch"""
    if not valid_files:
        return []
    
    results = []
    
    # Create progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Results container for real-time updates
    results_container = st.empty()
    
    total_files = len(valid_files)
    
    for i, uploaded_file in enumerate(valid_files):
        # Update progress
        progress = (i + 1) / total_files
        progress_bar.progress(progress)
        status_text.text(f"Processing {uploaded_file.name} ({i+1}/{total_files})")
        
        # Process single image
        result = process_single_image(uploaded_file, i, total_files)
        results.append(result)
        
        # Update results display in real-time
        with results_container.container():
            display_batch_progress(results, total_files)
        
        # Brief pause for UI responsiveness
        time.sleep(0.1)
    
    # Clear progress indicators
    progress_bar.empty()
    status_text.empty()
    results_container.empty()
    
    return results

def display_batch_progress(results: List[Dict], total_files: int):
    """Display real-time batch processing progress"""
    processed = len(results)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("📁 Processed", f"{processed}/{total_files}")
    
    with col2:
        compliant = sum(1 for r in results if r.get('compliance_status') == 'COMPLIANT')
        st.metric("✅ Compliant", compliant)
    
    with col3:
        errors = sum(1 for r in results if r.get('compliance_status') == 'ERROR')
        st.metric("❌ Errors", errors)

def display_batch_results(results: List[Dict[str, Any]], context: str = "default"):
    """Display comprehensive batch processing results"""
    if not results:
        return
    
    st.markdown("## 📊 Batch Processing Results")
    
    # Summary metrics
    total_files = len(results)
    successful = sum(1 for r in results if 'error' not in r)
    compliant = sum(1 for r in results if r.get('compliance_status') == 'COMPLIANT')
    errors = sum(1 for r in results if 'error' in r)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📁 Total Files", total_files)
    
    with col2:
        st.metric("✅ Successfully Processed", successful)
    
    with col3:
        st.metric("⚖️ Compliant Products", compliant)
    
    with col4:
        st.metric("❌ Processing Errors", errors)
    
    # Results table
    st.markdown("### 📋 Detailed Results")
    
    # Prepare data for table
    table_data = []
    for result in results:
        table_data.append({
            'Filename': result.get('filename', 'Unknown'),
            'Status': result.get('compliance_status', 'Unknown'),
            'Violations': len(result.get('violations', [])),
            'Processing Time': f"{result.get('processing_time', 0):.2f}s",
            'File Size': f"{result.get('file_size', 0) / 1024:.1f} KB",
            'Timestamp': result.get('timestamp', 'Unknown')
        })
    
    df = pd.DataFrame(table_data)
    
    # Interactive table with selection
    selected_indices = st.dataframe(
        df,
        width="stretch",
        selection_mode="multi-row",
        on_select="rerun",
        key=f"batch_results_dataframe_{context}"
    )
    
    # Detailed view for selected results
    if hasattr(selected_indices, 'selection') and selected_indices.selection.rows:
        st.markdown("### 🔍 Detailed View")
        
        for idx in selected_indices.selection.rows:
            if idx < len(results):
                display_single_result_detailed(results[idx])

def display_single_result_detailed(result: Dict[str, Any]):
    """Display detailed view of a single result"""
    filename = result.get('filename', 'Unknown')
    
    with st.expander(f"📄 {filename} - Detailed Results", expanded=True):
        
        # Error handling
        if 'error' in result:
            st.error(f"❌ Processing Error: {result['error']}")
            return
        
        # Result tabs
        tab1, tab2, tab3 = st.tabs(["📋 Structured Data", "⚖️ Compliance", "🔧 Technical"])
        
        with tab1:
            refined_data = result.get('refined_data', {})
            if refined_data:
                for field, value in refined_data.items():
                    col1, col2 = st.columns([1, 2])
                    with col1:
                        st.write(f"**{field.replace('_', ' ').title()}:**")
                    with col2:
                        if value:
                            st.write(f"`{value}`")
                        else:
                            st.write("*Not found*")
            else:
                st.warning("No structured data extracted")
        
        with tab2:
            violations = result.get('violations', [])
            compliance_status = result.get('compliance_status', 'UNKNOWN')
            
            if compliance_status == 'COMPLIANT':
                st.success("🎉 **FULLY COMPLIANT**")
            else:
                st.error(f"❌ **NON-COMPLIANT** - {len(violations)} violations")
                
                for i, violation in enumerate(violations, 1):
                    severity = violation.get('severity', 'Unknown').upper()
                    rule_id = violation.get('rule_id', 'UNKNOWN_RULE')
                    description = violation.get('description', 'No description')
                    
                    if severity == 'CRITICAL':
                        st.error(f"**{i}. {rule_id}** (Critical)")
                    elif severity == 'HIGH':
                        st.warning(f"**{i}. {rule_id}** (High)")
                    else:
                        st.info(f"**{i}. {rule_id}** (Medium)")
                    
                    st.write(f"   {description}")
        
        with tab3:
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Processing Information:**")
                st.write(f"- Processing Time: {result.get('processing_time', 0):.2f}s")
                st.write(f"- File Size: {result.get('file_size', 0) / 1024:.1f} KB")
                st.write(f"- Timestamp: {result.get('timestamp', 'Unknown')}")
            
            with col2:
                st.write("**Image Information:**")
                dimensions = result.get('image_dimensions', (0, 0))
                st.write(f"- Dimensions: {dimensions[0]} × {dimensions[1]} pixels")
                st.write(f"- Method: {result.get('method', 'Unknown')}")

def create_export_panel(results: List[Dict[str, Any]], context: str = "default"):
    """Create export functionality panel"""
    if not results:
        return
    
    st.markdown("### 📤 Export Results")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Export to Excel", width="stretch", key=f"export_excel_{context}"):
            excel_data = prepare_excel_export(results)
            st.download_button(
                label="⬇️ Download Excel Report",
                data=excel_data,
                file_name=f"compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{context}"
            )
    
    with col2:
        if st.button("📋 Export to JSON", width="stretch", key=f"export_json_{context}"):
            json_data = json.dumps(results, indent=2, ensure_ascii=False)
            st.download_button(
                label="⬇️ Download JSON Data",
                data=json_data,
                file_name=f"compliance_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key=f"download_json_{context}"
            )
    
    with col3:
        if st.button("📄 Generate PDF Report", width="stretch", key=f"export_pdf_{context}"):
            st.info("PDF report generation coming soon!")

def prepare_excel_export(results: List[Dict[str, Any]]) -> bytes:
    """Prepare Excel export data"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Summary data
    total_files = len(results)
    compliant = sum(1 for r in results if r.get('compliance_status') == 'COMPLIANT')
    non_compliant = total_files - compliant
    
    summary_data = [
        ["Legal Metrology Compliance Report", ""],
        ["Generated on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Files Processed", total_files],
        ["Compliant Products", compliant],
        ["Non-Compliant Products", non_compliant],
        ["Compliance Rate", f"{(compliant/total_files*100):.1f}%" if total_files > 0 else "0%"]
    ]
    
    for row_idx, (key, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=key)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    # Detailed results sheet
    detail_ws = wb.create_sheet("Detailed Results")
    
    # Headers
    headers = ["Filename", "Compliance Status", "Violations Count", "Processing Time (s)", "File Size (KB)", "Timestamp"]
    for col_idx, header in enumerate(headers, 1):
        cell = detail_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    for row_idx, result in enumerate(results, 2):
        detail_ws.cell(row=row_idx, column=1, value=result.get('filename', 'Unknown'))
        detail_ws.cell(row=row_idx, column=2, value=result.get('compliance_status', 'Unknown'))
        detail_ws.cell(row=row_idx, column=3, value=len(result.get('violations', [])))
        detail_ws.cell(row=row_idx, column=4, value=result.get('processing_time', 0))
        detail_ws.cell(row=row_idx, column=5, value=result.get('file_size', 0) / 1024)
        detail_ws.cell(row=row_idx, column=6, value=result.get('timestamp', 'Unknown'))
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def main():
    """Main function for upload page"""
    # Load custom CSS
    load_upload_css()
    
    # Initialize session state
    initialize_upload_session()
    
    # Page header
    st.markdown("""
    # 📂 Upload & Process Images
    
    Upload multiple product images for batch compliance validation.
    Drag and drop images or use the upload button to get started.
    """)
    
    # Sidebar controls
    with st.sidebar:
        st.markdown("### 📊 Processing Options")
        
        # Batch processing settings
        auto_process = st.checkbox(
            "Auto-process on upload",
            value=True,
            help="Automatically process images when uploaded"
        )
        
        confidence_threshold = st.slider(
            "OCR Confidence Threshold:",
            min_value=0.1,
            max_value=1.0,
            value=0.4,
            step=0.1
        )
        
        save_results = st.checkbox(
            "Save results to session",
            value=True,
            help="Keep results for comparison and export"
        )
        
        st.markdown("---")
        
        # Session statistics
        st.markdown("### 📈 Session Stats")
        
        total_processed = len(st.session_state.batch_results)
        st.metric("📁 Files Processed", total_processed)
        
        if st.session_state.batch_results:
            compliant_count = sum(1 for r in st.session_state.batch_results 
                                if r.get('compliance_status') == 'COMPLIANT')
            st.metric("✅ Compliant Rate", f"{(compliant_count/total_processed*100):.1f}%")
    
    # Main upload interface
    uploaded_files = create_file_upload_interface()
    
    # Display uploaded files
    if uploaded_files:
        valid_files = display_uploaded_files(uploaded_files)
        
        if valid_files:
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                st.success(f"✅ {len(valid_files)} valid files ready for processing")
            
            with col2:
                process_button = st.button(
                    "🚀 Process All Images",
                    type="primary",
                    width="stretch",
                    disabled=len(valid_files) == 0,
                    key="process_all_images"
                )
            
            with col3:
                clear_button = st.button(
                    "🗑️ Clear Selection",
                    width="stretch",
                    key="clear_current_selection"
                )
                
                if clear_button:
                    # Reset all processing state
                    st.session_state.current_batch_results = []
                    st.session_state.files_processed = set()
                    st.session_state.processing_completed = False
                    st.rerun()
            
            # Create unique file identifiers for current upload
            current_file_ids = {f"{f.name}_{f.size}" for f in valid_files}
            
            # Check if we need to process (only if files have changed or manual processing requested)
            should_process = (
                process_button or 
                (auto_process and valid_files and not st.session_state.processing_completed) or
                (auto_process and valid_files and current_file_ids != st.session_state.files_processed)
            )
            
            # Process images
            if should_process:
                with st.spinner("Processing images through compliance pipeline..."):
                    batch_results = process_batch_images(valid_files)
                    
                    # Update session state
                    st.session_state.current_batch_results = batch_results
                    st.session_state.files_processed = current_file_ids
                    st.session_state.processing_completed = True
                    
                    if save_results:
                        st.session_state.batch_results.extend(batch_results)
            
            # Display current results (from session state to avoid reprocessing)
            if st.session_state.current_batch_results:
                st.success(f"✅ Successfully processed {len(st.session_state.current_batch_results)} images!")
                display_batch_results(st.session_state.current_batch_results, "current_batch")
                
                # Export panel
                create_export_panel(st.session_state.current_batch_results, "current_batch")
    
    # Display session results if available
    if st.session_state.batch_results:
        st.markdown("---")
        st.markdown("## 📚 Session History")
        
        with st.expander(f"View All Results ({len(st.session_state.batch_results)} files)", expanded=False):
            display_batch_results(st.session_state.batch_results, "session_history")
            create_export_panel(st.session_state.batch_results, "session_history")
        
        # Clear session history
        if st.button("🗑️ Clear Session History", key="clear_session_history"):
            st.session_state.batch_results = []
            st.rerun()
    
    # Navigation
    st.markdown("---")
    nav_col1, nav_col2, nav_col3 = st.columns(3)
    
    with nav_col1:
        if st.button("🏠 Back to Dashboard", width="stretch", key="nav_dashboard_upload"):
            st.switch_page("streamlit_app.py")
    
    with nav_col2:
        if st.button("📷 Live Camera", width="stretch", key="nav_camera_upload"):
            st.switch_page("pages/01_📷_Live_Camera.py")
    
    with nav_col3:
        if st.button("📊 Batch Processing", width="stretch", key="nav_batch_upload"):
            st.switch_page("pages/03_📊_Batch_Process.py")

if __name__ == "__main__":
    main()