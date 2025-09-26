"""
Report Generator - Legal Metrology OCR Pipeline

Professional report generation for compliance validation results.
Supports PDF, Excel, and HTML report formats with customizable templates.
"""

import streamlit as st
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
import io
import base64
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ReportGenerator:
    """Professional report generator for compliance validation results"""
    
    def __init__(self):
        self.report_templates = {
            'compliance': self.generate_compliance_report,
            'batch_summary': self.generate_batch_summary_report,
            'analytics': self.generate_analytics_report,
            'detailed': self.generate_detailed_report
        }
    
    def generate_report(self, data: Dict[str, Any], report_type: str = 'compliance', 
                       format: str = 'pdf') -> bytes:
        """Generate report in specified format"""
        
        if report_type not in self.report_templates:
            raise ValueError(f"Unsupported report type: {report_type}")
        
        if format == 'pdf':
            return self.generate_pdf_report(data, report_type)
        elif format == 'excel':
            return self.generate_excel_report(data, report_type)
        elif format == 'html':
            return self.generate_html_report(data, report_type)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def generate_pdf_report(self, data: Dict[str, Any], report_type: str) -> bytes:
        """Generate PDF report using ReportLab"""
        
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#1f4037'),
            alignment=1  # Center alignment
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#1f4037'),
            borderWidth=1,
            borderColor=colors.HexColor('#1f4037'),
            borderPadding=5
        )
        
        # Build report content
        content = []
        
        # Title
        content.append(Paragraph("Legal Metrology OCR Compliance Report", title_style))
        content.append(Spacer(1, 20))
        
        # Report metadata
        metadata_data = [
            ['Generated On:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['Report Type:', report_type.title()],
            ['System:', 'Legal Metrology OCR Pipeline v2.0']
        ]
        
        metadata_table = Table(metadata_data)
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        content.append(metadata_table)
        content.append(Spacer(1, 20))
        
        # Generate report content based on type
        report_content = self.report_templates[report_type](data, 'pdf')
        content.extend(report_content)
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def generate_excel_report(self, data: Dict[str, Any], report_type: str) -> bytes:
        """Generate Excel report using openpyxl"""
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate report content based on type
        self.report_templates[report_type](data, 'excel', wb)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def generate_html_report(self, data: Dict[str, Any], report_type: str) -> str:
        """Generate HTML report"""
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Legal Metrology OCR Compliance Report</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    max-width: 1200px;
                    margin: 0 auto;
                    padding: 20px;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    background: linear-gradient(135deg, #1f4037, #99f2c8);
                    color: white;
                    padding: 2rem;
                    border-radius: 10px;
                    margin-bottom: 2rem;
                }}
                .section {{
                    margin: 2rem 0;
                    padding: 1rem;
                    border: 1px solid #ddd;
                    border-radius: 8px;
                }}
                .compliant {{
                    background-color: #d4edda;
                    color: #155724;
                    border-color: #c3e6cb;
                }}
                .non-compliant {{
                    background-color: #f8d7da;
                    color: #721c24;
                    border-color: #f5c6cb;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 1rem 0;
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }}
                th {{
                    background-color: #1f4037;
                    color: white;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Legal Metrology OCR Compliance Report</h1>
                <p>Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            </div>
        """
        
        # Generate report content based on type
        report_content = self.report_templates[report_type](data, 'html')
        html_content += report_content
        
        html_content += """
        </body>
        </html>
        """
        
        return html_content
    
    def generate_compliance_report(self, data: Dict[str, Any], format: str, wb=None) -> Any:
        """Generate compliance-focused report"""
        
        if format == 'pdf':
            content = []
            styles = getSampleStyleSheet()
            
            # Compliance status
            violations = data.get('violations', [])
            compliance_status = data.get('compliance_status', 'UNKNOWN')
            
            content.append(Paragraph("Compliance Assessment", styles['Heading2']))
            
            status_text = f"Status: {compliance_status}"
            if compliance_status == 'COMPLIANT':
                status_text += " ✓"
            else:
                status_text += f" - {len(violations)} violations found"
            
            content.append(Paragraph(status_text, styles['Normal']))
            content.append(Spacer(1, 12))
            
            # Violations table
            if violations:
                content.append(Paragraph("Violations Details", styles['Heading3']))
                
                violation_data = [['Rule ID', 'Severity', 'Description']]
                for violation in violations:
                    violation_data.append([
                        violation.get('rule_id', 'UNKNOWN'),
                        violation.get('severity', 'MEDIUM'),
                        violation.get('description', 'No description')[:60] + '...'
                    ])
                
                violation_table = Table(violation_data)
                violation_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4037')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                content.append(violation_table)
            
            return content
        
        elif format == 'excel':
            # Summary sheet
            summary_ws = wb.create_sheet("Compliance Summary")
            
            # Header
            summary_ws['A1'] = 'Legal Metrology Compliance Report'
            summary_ws['A1'].font = Font(size=16, bold=True)
            
            # Compliance data
            violations = data.get('violations', [])
            compliance_status = data.get('compliance_status', 'UNKNOWN')
            
            summary_ws['A3'] = 'Compliance Status'
            summary_ws['B3'] = compliance_status
            
            summary_ws['A4'] = 'Total Violations'
            summary_ws['B4'] = len(violations)
            
            summary_ws['A5'] = 'Report Generated'
            summary_ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Violations detail sheet
            if violations:
                violations_ws = wb.create_sheet("Violations Detail")
                
                headers = ['Rule ID', 'Severity', 'Field', 'Description']
                for col, header in enumerate(headers, 1):
                    cell = violations_ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for row, violation in enumerate(violations, 2):
                    violations_ws.cell(row=row, column=1, value=violation.get('rule_id', 'UNKNOWN'))
                    violations_ws.cell(row=row, column=2, value=violation.get('severity', 'MEDIUM'))
                    violations_ws.cell(row=row, column=3, value=violation.get('violating_field', 'N/A'))
                    violations_ws.cell(row=row, column=4, value=violation.get('description', 'No description'))
        
        elif format == 'html':
            violations = data.get('violations', [])
            compliance_status = data.get('compliance_status', 'UNKNOWN')
            
            html = f"""
            <div class="section {'compliant' if compliance_status == 'COMPLIANT' else 'non-compliant'}">
                <h2>Compliance Assessment</h2>
                <p><strong>Status:</strong> {compliance_status}</p>
                <p><strong>Violations Found:</strong> {len(violations)}</p>
            </div>
            """
            
            if violations:
                html += """
                <div class="section">
                    <h3>Violation Details</h3>
                    <table>
                        <tr>
                            <th>Rule ID</th>
                            <th>Severity</th>
                            <th>Field</th>
                            <th>Description</th>
                        </tr>
                """
                
                for violation in violations:
                    html += f"""
                        <tr>
                            <td>{violation.get('rule_id', 'UNKNOWN')}</td>
                            <td>{violation.get('severity', 'MEDIUM')}</td>
                            <td>{violation.get('violating_field', 'N/A')}</td>
                            <td>{violation.get('description', 'No description')}</td>
                        </tr>
                    """
                
                html += """
                    </table>
                </div>
                """
            
            return html
    
    def generate_batch_summary_report(self, data: List[Dict[str, Any]], format: str, wb=None) -> Any:
        """Generate batch processing summary report"""
        
        if not isinstance(data, list):
            data = [data]
        
        total_files = len(data)
        compliant_files = len([d for d in data if d.get('compliance_status') == 'COMPLIANT'])
        avg_processing_time = sum(d.get('processing_time', 0) for d in data) / total_files if total_files > 0 else 0
        
        if format == 'pdf':
            content = []
            styles = getSampleStyleSheet()
            
            content.append(Paragraph("Batch Processing Summary", styles['Heading2']))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Files Processed', str(total_files)],
                ['Compliant Files', str(compliant_files)],
                ['Compliance Rate', f"{(compliant_files/total_files*100):.1f}%" if total_files > 0 else "0%"],
                ['Average Processing Time', f"{avg_processing_time:.2f}s"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4037')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            content.append(summary_table)
            
            return content
        
        elif format == 'excel':
            # Summary sheet
            summary_ws = wb.create_sheet("Batch Summary")
            
            summary_ws['A1'] = 'Batch Processing Summary'
            summary_ws['A1'].font = Font(size=16, bold=True)
            
            summary_ws['A3'] = 'Total Files'
            summary_ws['B3'] = total_files
            
            summary_ws['A4'] = 'Compliant Files'
            summary_ws['B4'] = compliant_files
            
            summary_ws['A5'] = 'Compliance Rate'
            summary_ws['B5'] = f"{(compliant_files/total_files*100):.1f}%" if total_files > 0 else "0%"
            
            summary_ws['A6'] = 'Avg Processing Time'
            summary_ws['B6'] = f"{avg_processing_time:.2f}s"
            
            # Detailed results sheet
            if PANDAS_AVAILABLE:
                details_ws = wb.create_sheet("Detailed Results")
                
                # Prepare data
                details_data = []
                for result in data:
                    details_data.append({
                        'Filename': result.get('filename', 'Unknown'),
                        'Compliance Status': result.get('compliance_status', 'UNKNOWN'),
                        'Violations': len(result.get('violations', [])),
                        'Processing Time': result.get('processing_time', 0),
                        'Timestamp': result.get('timestamp', 'Unknown')
                    })
                
                df = pd.DataFrame(details_data)
                
                # Write to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    details_ws.append(r)
                
                # Style headers
                for cell in details_ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        elif format == 'html':
            html = f"""
            <div class="section">
                <h2>Batch Processing Summary</h2>
                <table>
                    <tr><th>Metric</th><th>Value</th></tr>
                    <tr><td>Total Files Processed</td><td>{total_files}</td></tr>
                    <tr><td>Compliant Files</td><td>{compliant_files}</td></tr>
                    <tr><td>Compliance Rate</td><td>{(compliant_files/total_files*100):.1f}%</td></tr>
                    <tr><td>Average Processing Time</td><td>{avg_processing_time:.2f}s</td></tr>
                </table>
            </div>
            """
            
            return html
    
    def generate_analytics_report(self, data: Dict[str, Any], format: str, wb=None) -> Any:
        """Generate analytics and insights report"""
        # Implementation for analytics report
        pass
    
    def generate_detailed_report(self, data: Dict[str, Any], format: str, wb=None) -> Any:
        """Generate comprehensive detailed report"""
        # Implementation for detailed report
        pass


def create_report_download_button(data: Dict[str, Any], report_type: str = 'compliance', 
                                format: str = 'pdf') -> None:
    """Create a download button for generated reports"""
    
    try:
        generator = ReportGenerator()
        report_data = generator.generate_report(data, report_type, format)
        
        filename = f"compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format == 'pdf':
            mime_type = "application/pdf"
            file_ext = "pdf"
        elif format == 'excel':
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_ext = "xlsx"
        elif format == 'html':
            mime_type = "text/html"
            file_ext = "html"
        else:
            st.error(f"Unsupported format: {format}")
            return
        
        st.download_button(
            label=f"📥 Download {format.upper()} Report",
            data=report_data,
            file_name=f"{filename}.{file_ext}",
            mime=mime_type,
            use_container_width=True
        )
        
    except ImportError as e:
        st.error(f"Required library not available: {str(e)}")
        st.info("Install additional dependencies for report generation")
    
    except Exception as e:
        st.error(f"Failed to generate report: {str(e)}")


def create_report_preview(data: Dict[str, Any], report_type: str = 'compliance') -> None:
    """Create a preview of the report content"""
    
    st.markdown("### 📄 Report Preview")
    
    if report_type == 'compliance':
        # Compliance report preview
        violations = data.get('violations', [])
        compliance_status = data.get('compliance_status', 'UNKNOWN')
        
        st.markdown(f"**Compliance Status:** {compliance_status}")
        st.markdown(f"**Violations Found:** {len(violations)}")
        st.markdown(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        if violations:
            st.markdown("**Sample Violations:**")
            for i, violation in enumerate(violations[:3]):  # Show first 3
                st.write(f"{i+1}. {violation.get('rule_id', 'UNKNOWN')}: {violation.get('description', 'No description')}")
    
    else:
        st.info(f"Preview not available for {report_type} reports")


# Utility functions for report customization
def get_company_branding() -> Dict[str, str]:
    """Get company branding information"""
    return {
        'company_name': 'Legal Metrology OCR Pipeline',
        'logo_path': Path(__file__).parent.parent / 'assets' / 'logo.png',
        'primary_color': '#1f4037',
        'secondary_color': '#99f2c8'
    }

def format_violation_for_report(violation: Dict[str, Any]) -> str:
    """Format violation for report display"""
    rule_id = violation.get('rule_id', 'UNKNOWN')
    severity = violation.get('severity', 'MEDIUM')
    description = violation.get('description', 'No description available')
    
    return f"[{severity}] {rule_id}: {description}"